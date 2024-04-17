# Trabalhando com openpyxl
# Criando e adicionando itens a planilha
from pathlib import Path

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

ROOT_FOLDER = Path(__file__).parent
WORKBOOK_PATH = ROOT_FOLDER / 'workbook.xlsx'

workbook = Workbook()
# worksheet: Worksheet = workbook.active

# nome da minha planilha
sheet_name = 'Minha Planilha'
# criamos a planilha
workbook.create_sheet(sheet_name)
# selecionou a planilha
worksheet: Worksheet = workbook[sheet_name]

# Remover uma planilha
workbook.remove(workbook['Sheet'])

# criando os cabeçalhos
worksheet.cell(1, 1, 'Nome')
worksheet.cell(1, 2, 'Idade')
worksheet.cell(1, 3, 'Nota')

students = [
    # nome   idade   nota
    ['João',  14, 5.5],
    ['Maria', 13, 9.7],
    ['Luiz',  15, 8.8],
    ['Alberto', 16, 10],
]

# for i, student_row in enumerate(studentes, start=2):
#     for j, student_column in enumerate(student_row, start=1):
#         worksheet.cell(i, j, student_column)

for student in students:
    worksheet.append(student)

workbook.save(WORKBOOK_PATH)
