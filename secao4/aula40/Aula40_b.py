# Lendo e alterar dados da planilha
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet

ROOT_FOLDER = Path(__file__).parent
WORKBOOK_PATH = ROOT_FOLDER / 'workbook.xlsx'

# carregando arquivo da planilha
workbook: Workbook = load_workbook(WORKBOOK_PATH)

# nome da minha planilha
sheet_name = 'Minha Planilha'

# selecionou a planilha
worksheet: Worksheet = workbook[sheet_name]

row: tuple[Cell]
for row in worksheet.iter_rows(min_row=2):
    for cell in row:
        print(cell.value, end='\t')
    print()

# workbook.save(WORKBOOK_PATH)
